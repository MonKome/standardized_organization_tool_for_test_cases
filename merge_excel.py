import configparser
import difflib
import logging
from pathlib import Path
import pandas as pd
from openpyxl.styles import Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl import Workbook,load_workbook
from openpyxl.utils import get_column_letter, range_boundaries

def setup_logger(log_file: str = "merge_log.log"):
    """设置日志记录器"""
    logger = logging.getLogger()
    logger.setLevel(logging.DEBUG)
    if logger.hasHandlers():
        logger.handlers.clear()

    formatter = logging.Formatter(
        "%(asctime)s - %(levelname)s - %(message)s", datefmt="%Y-%m-%d %H:%M:%S"
    )

    console_handler = logging.StreamHandler()
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(formatter)

    file_handler = logging.FileHandler(log_file, mode="a", encoding="utf-8")
    file_handler.setLevel(logging.DEBUG)
    file_handler.setFormatter(formatter)

    logger.addHandler(console_handler)
    logger.addHandler(file_handler)


def create_synonym_map(config_parser: configparser.ConfigParser) -> dict:
    """
    从配置文件的 [column_mapping] 部分构建近义词映射字典。
    """
    synonym_map = {}
    if config_parser.has_section("column_mapping"):
        for standard_col in config_parser.options("column_mapping"):
            if standard_col != "template_columns":
                synonyms_str = config_parser.get("column_mapping", standard_col)
                synonyms = [syn.strip() for syn in synonyms_str.split(",")]
                synonym_map[standard_col] = synonyms
                logging.debug(f"加载近义词映射: '{standard_col}' <- {synonyms}")
    else:
        logging.warning("[column_mapping] 部分未在配置文件中找到。将仅使用精确匹配。")
    return synonym_map


def map_columns_to_template(
    file_columns: list, template_columns: list, synonym_map: dict
) -> dict:
    """
    将文件的实际列名映射到标准模板列名。
    三级匹配策略：精确匹配 → 近义词匹配 → 模糊匹配
    """
    column_mapping = {}
    for file_col in file_columns:
        mapped_col = None
        # 1. 精确匹配
        if file_col in template_columns:
            mapped_col = file_col
        else:
            # 2. 检查近义词映射
            for std_col, synonyms in synonym_map.items():
                if file_col in synonyms:
                    mapped_col = std_col
                    break

            # 3. 模糊匹配 (difflib)
            if not mapped_col:
                all_possible_names = template_columns[:]
                for std_col, synonyms in synonym_map.items():
                    all_possible_names.extend(synonyms)

                closest_match = difflib.get_close_matches(
                    file_col, all_possible_names, n=1, cutoff=0.6
                )
                if closest_match:
                    match = closest_match[0]
                    if match in template_columns:
                        mapped_col = match
                    else:
                        for std_col, synonyms in synonym_map.items():
                            if match in synonyms:
                                mapped_col = std_col
                                break

        if mapped_col:
            column_mapping[file_col] = mapped_col
            logging.debug(f"  映射列: '{file_col}' -> '{mapped_col}'")
        else:
            logging.warning(f"  无法识别的列 '{file_col}' 将被忽略。")
    return column_mapping


def parse_sheet_names(raw_value: str) -> list:
    """
    解析 sheet_name 配置值，支持逗号（半角/全角）、顿号、空格等多种分隔符。
    """
    import re
    names = re.split(r"[，,、\s]+", raw_value.strip())
    return [name for name in names if name]


def validate_config(config_parser: configparser.ConfigParser) -> bool:
    """验证配置文件的内容。"""
    required_sections = ["paths", "settings", "column_mapping"]
    required_options = {
        "paths": ["input_folder", "output_path"],
        "settings": ["sheet_name"],
        "column_mapping": ["template_columns"],
    }

    is_valid = True
    for section in required_sections:
        if not config_parser.has_section(section):
            logging.error(f"配置文件中缺少必要的节 '[{section}]'。")
            is_valid = False
            continue
        for option in required_options.get(section, []):
            if not config_parser.has_option(section, option):
                logging.error(
                    f"配置文件中节 '[{section}]' 下缺少必要的选项 '{option}'。"
                )
                is_valid = False

    if is_valid:
        try:
            input_folder = config_parser.get("paths", "input_folder")
            output_path = config_parser.get("paths", "output_path")

            if not Path(input_folder).is_dir():
                logging.error(f"配置中的输入文件夹路径无效或不存在: '{input_folder}'")
                is_valid = False

            output_dir = Path(output_path).parent
            try:
                output_dir.mkdir(parents=True, exist_ok=True)
            except Exception as e:
                logging.error(
                    f"无法创建或访问输出路径的目录: '{output_dir}'. 错误: {e}"
                )
                is_valid = False

            # 验证 sheet_name 不为空且至少有一项
            raw_sheet = config_parser.get("settings", "sheet_name")
            sheet_names = parse_sheet_names(raw_sheet)
            if not sheet_names:
                logging.error("配置中的 'sheet_name' 不能为空。")
                is_valid = False

        except configparser.Error as e:
            logging.error(f"读取配置项时发生错误: {e}")
            is_valid = False

    return is_valid


def load_config_from_ini(config_file: str) -> tuple:
    """
    从INI文件加载配置。
    Returns:
        tuple: (template_columns, synonym_map, sheet_names, source_column_name)
    """
    logging.info(f"正在读取配置文件: {config_file}")
    config_parser = configparser.ConfigParser()
    try:
        with open(config_file, "r", encoding="utf-8") as f:
            config_parser.read_file(f)
    except FileNotFoundError:
        logging.error(f"配置文件未找到: {config_file}")
        raise
    except Exception as e:
        logging.error(f"读取配置文件时发生错误: {e}")
        raise

    if not validate_config(config_parser):
        raise ValueError("配置文件验证失败，请检查 config.ini 文件。")

    # 解析模板列
    template_cols_str = config_parser.get("column_mapping", "template_columns")
    template_columns = [col.strip() for col in template_cols_str.split(",")]
    logging.info(f"加载的标准模板列: {template_columns}")

    # 构建近义词映射
    synonym_map = create_synonym_map(config_parser)

    # 解析目标Sheet名称列表
    raw_sheet = config_parser.get("settings", "sheet_name")
    sheet_names = parse_sheet_names(raw_sheet)
    logging.info(f"目标Sheet名称列表: {sheet_names}")

    # 来源文件名列名
    source_column_name = config_parser.get("settings", "source_column_name", fallback="来源文件")
    logging.info(f"来源文件列名: '{source_column_name}'")

    return template_columns, synonym_map, sheet_names, source_column_name


def get_all_files(folder_path: str, extension: str = ".xlsx") -> list:
    """获取指定文件夹内所有指定扩展名的文件"""
    logging.info(f"正在扫描文件夹 '{folder_path}' 中的Excel文件...")
    p = Path(folder_path)
    if not p.exists() or not p.is_dir():
        logging.warning(f"输入文件夹路径无效或不存在: {folder_path}")
        return []
    files = [f for f in p.rglob(f"*{extension}") if f.is_file()]
    logging.info(f"在 '{folder_path}' 中找到 {len(files)} 个Excel文件 (.xlsx)")
    return sorted(files)


def apply_styling(ws, df):
    """
    为工作表应用样式：细边框 + 自动换行 + 垂直居中 + 自适应列宽。
    """
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    centered_alignment = Alignment(wrap_text=True, vertical="center")

    for row in ws.iter_rows(
        min_row=1, max_row=df.shape[0] + 1, min_col=1, max_col=df.shape[1]
    ):
        for cell in row:
            cell.border = thin_border
            cell.alignment = centered_alignment

    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width


def process_file_sheet(file_path, sheet_name, template_columns, synonym_map,
                       source_column_name, source_sheet_name=None):
    """
    从单个文件的单个Sheet中读取数据、映射列、添加来源信息。
    Args:
        file_path: Excel文件路径
        sheet_name: 要读取的Sheet名称
        template_columns: 标准模板列名列表
        synonym_map: 近义词映射字典
        source_column_name: "来源文件"列的列名
        source_sheet_name: "来源Sheet"列的列名（多Sheet时使用，单Sheet时为None）
    Returns:
        DataFrame 或 None（读取失败时）
    """
    try:
        wb_check = load_workbook(file_path, read_only=True, data_only=False)
        if sheet_name not in wb_check.sheetnames:
            wb_check.close()
            return pd.DataFrame(), []
        df = pd.read_excel(file_path, sheet_name=sheet_name)
        merged_ranges = []
        # 2. 新增：用 openpyxl 读取合并单元格信息
        wb_source = load_workbook(file_path, read_only=False, data_only=False)
        if sheet_name in wb_source.sheetnames:
            ws_source = wb_source[sheet_name]
            # 获取合并的区域列表，例如 ['A1:B1', 'C3:D5']
            merged_ranges = [str(range) for range in ws_source.merged_cells.ranges]
        wb_source.close()
    except ValueError:
        logging.warning(f"  文件 '{file_path.name}' 中不存在Sheet '{sheet_name}'，跳过。")
        return None,[]
    except Exception as e:
        logging.error(f"  读取文件 '{file_path.name}' 的Sheet '{sheet_name}' 失败: {e}")
        return None,[]
    if df.empty:
        logging.warning(f"  文件 '{file_path.name}' 中Sheet '{sheet_name}' 为空，跳过。")
        return None,[]
    logging.debug(f"  原始列: {list(df.columns)}")

    # 列映射
    col_mapping = map_columns_to_template(df.columns.tolist(), template_columns, synonym_map)

    # 选择目标列并重命名
    relevant_cols = [
        f_col for f_col, t_col in col_mapping.items()
        if t_col in template_columns
    ]
    if not relevant_cols:
        logging.warning( f"  文件 '{file_path.name}' Sheet '{sheet_name}' 没有匹配到任何模板列，跳过。")
        return None,[]

    df_subset = df[relevant_cols].copy()
    rename_dict = {
        f_col: t_col for f_col, t_col in col_mapping.items()
        if t_col in template_columns
    }
    df_subset.rename(columns=rename_dict, inplace=True)

    # 添加来源信息
    df_subset[source_column_name] = file_path.name
    if source_sheet_name is not None:
        df_subset[source_sheet_name] = sheet_name

    df_subset.reset_index(drop=True, inplace=True)
    logging.info(f"  [{file_path.name}] Sheet '{sheet_name}' -> {df_subset.shape[0]} 行")
    return df_subset,merged_ranges


def merge_sheets_with_config(config_file: str):
    """
    根据配置文件合并Excel工作表，并应用样式。
    """
    output_path = None
    try:
        # ---------- 加载配置 ----------
        template_columns, synonym_map, sheet_names, source_column_name = load_config_from_ini(config_file)

        config_parser = configparser.ConfigParser()
        config_parser.read(config_file, encoding="utf-8")
        input_folder = config_parser.get("paths", "input_folder")
        output_path = config_parser.get("paths", "output_path")

        logging.info(f"确认输出目录存在: {Path(output_path).parent}")

        # ---------- 扫描文件 ----------
        excel_files = get_all_files(input_folder)
        if not excel_files:
            logging.warning(f"在文件夹 '{input_folder}' 中没有找到任何Excel文件 (.xlsx)，程序即将退出。")
            return

        logging.info(f"准备处理的文件列表: {[f.name for f in excel_files]}")

        # ---------- 处理数据 ----------
        is_multi_sheet = len(sheet_names) > 1
        sheet_label = "来源Sheet" if is_multi_sheet else None
        all_dataframes = []
        all_merged_ranges = []  # 新增：用于存储所有文件的合并范围

        logging.info("开始读取和处理各个文件...")
        for file_path in excel_files:
            for sheet_name in sheet_names:
                df_subset, merged_ranges = process_file_sheet(file_path, sheet_name, template_columns, synonym_map,source_column_name, sheet_label)
                if df_subset is not None and df_subset.shape[0]>0:
                    all_dataframes.append(df_subset)
                    # 新增：保存合并范围
                    all_merged_ranges.append(merged_ranges)
        if not all_dataframes:
            logging.critical("没有成功读取到任何数据，无法继续合并操作，程序退出。")
            return

        logging.info(f"成功从 {len(all_dataframes)} 个Sheet/文件组合中读取到数据。")

        # ---------- 合并 ----------
        logging.info("正在合并所有数据...")
        extra_cols = [source_column_name]
        if is_multi_sheet and sheet_label:
            extra_cols = [source_column_name, sheet_label]
        final_columns_order = template_columns + extra_cols

        final_df = pd.concat(all_dataframes, ignore_index=True, sort=False)
        final_df = final_df.reindex(columns=final_columns_order)

        logging.info(f"合并后的总数据量: {final_df.shape[0]} 行, {final_df.shape[1]} 列")
        logging.debug(f"最终列顺序: {list(final_df.columns)}")

        # ---------- 写入 ----------
        wb = Workbook()
        ws = wb.active
        # 输出Sheet名称：单Sheet沿用原名，多Sheet使用通用名称
        ws.title = sheet_names[0] if len(sheet_names) == 1 else "合并汇总"
        # 写入表头
        for c_idx, value in enumerate(final_df.columns, 1):
            ws.cell(row=1, column=c_idx, value=value)
        # 写入数据
        for r_idx, row in enumerate(dataframe_to_rows(final_df, index=False, header=False), 2):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)
        # 应用合并单元格逻辑
        logging.info("正在应用合并单元格范围...")

        temp_offset = 0  # 重置偏移量，指向数据开始的上一行（即表头行）
        df_index = 0

        for ranges in all_merged_ranges:
            df_rows = all_dataframes[df_index].shape[0]  # 当前这个文件有多少行数据
            for range_str in ranges:
                try:
                    min_col, min_row, max_col, max_row = [int(x) for x in range_boundaries(range_str)]
                    # 计算新的行号（原行号 + 偏移量）
                    new_min_row = min_row + temp_offset
                    new_max_row = max_row + temp_offset

                    ws.merge_cells(start_row=new_min_row, start_column=min_col, end_row=new_max_row,
                                   end_column=max_col)
                    # 3. 【新增】将列号(数字)转回字母(如 1 -> A)，并拼接成 "A1:B2" 格式
                    # get_column_letter 用于将数字列号转为字母
                    start_cell = f"{get_column_letter(min_col)}{new_min_row}"
                    end_cell = f"{get_column_letter(max_col)}{new_max_row}"
                    formatted_range = f"[{start_cell}:{end_cell}]"
                    # 打印调试信息
                    logging.debug(f"原范围: {range_str} -> 偏移后: {formatted_range}")

                except Exception as e:
                    logging.warning(f"合并单元格失败 {range_str}: {e}")

            # 更新偏移量，为下一个文件做准备
            temp_offset += df_rows
            df_index += 1

        logging.info("正在为工作表应用样式...")
        # 应用其他样式（边框、宽度等）
        apply_styling(ws, final_df)

        logging.info(f"正在将合并后的数据（含样式）写入文件: {output_path}")
        wb.save(output_path)

        logging.info(f"任务成功完成! 已生成文件: {output_path}")

    except PermissionError:
        logging.critical(f"权限不足，无法写入文件或目录: {output_path}")
    except Exception as e:
        logging.critical(
            f"程序执行过程中发生致命错误: {type(e).__name__} - {e}", exc_info=True
        )
        raise


if __name__ == "__main__":
    setup_logger("merge_log.log")
    logging.info("--- Excel文件合并脚本启动 ---")
    merge_sheets_with_config("config.ini")
    logging.info("--- Excel文件合并脚本结束 ---")
